#!/usr/bin/env python3
"""BrowserForensics.py - multi-source browser history & download triage.

Accepts one or more Chromium-family `History` files (Chrome/Edge/Brave/...)
and/or Firefox `places.sqlite` files -- or directories to search for them --
and writes a single Excel workbook:

    Summary             source paths, SHA-256, WAL sidecars, record counts
    Combined Timeline   every record from every source, newest first
    <one sheet per source DB, auto-named e.g. Chrome-Default, Firefox-Default>

Record types per sheet: Visit, Download, Search Term (Chromium omnibox
searches), Typed Input (Firefox awesome-bar input history).

Falls back to per-source CSVs if openpyxl is unavailable, or if the output
filename ends in .csv.

Usage:
    python BrowserForensics.py <db_or_dir> [<db_or_dir> ...] [-o out.xlsx]

Examples:
    python BrowserForensics.py "C:\\Users\\jdoe\\AppData\\Local\\Google\\Chrome\\User Data\\Default\\History" ^
                               "C:\\Users\\jdoe\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\ab12cd.default-release\\places.sqlite" ^
                               -o case1234.xlsx
    python BrowserForensics.py D:\\Collections\\host01\\browsers -o case1234.xlsx
"""

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import sqlite3
import sys
import tempfile
from datetime import datetime, timezone

# Chromium timestamps: microseconds since 1601-01-01 UTC (WebKit/FILETIME-style)
CHROMIUM_EPOCH_OFFSET_US = 11_644_473_600_000_000
EXCEL_MAX_ROWS = 1_048_576

FIELDNAMES = [
    "Type",
    "Visit ID",
    "URL",
    "Title",
    "Search Term",
    "Visit Count",
    "Visit Time (UTC)",
    "Last Visit (UTC)",
    "End Time (UTC)",
    "Visit Time Raw",
    "Last Visit Raw",
    "Visit Type",
    "Visit Type Raw",
    "Transition Qualifiers",
    "Transition Raw",
    "From Visit ID",
    "Opener Visit ID",
    "Segment ID",
    "Visit Duration (s)",
    "Download Path",
    "Download URL",
    "Referrer",
    "MIME Type",
    "Download State",
    "Danger Type",
    "Interrupt Reason",
    "Download Size (bytes)",
    "Received Bytes",
]

COLUMN_WIDTHS = {
    "Type": 12, "Visit ID": 9, "URL": 60, "Title": 40, "Search Term": 28,
    "Visit Time (UTC)": 20, "Last Visit (UTC)": 20, "End Time (UTC)": 20,
    "Visit Time Raw": 18, "Last Visit Raw": 18, "Visit Type": 16,
    "Transition Qualifiers": 30, "Download Path": 45, "Download URL": 50,
    "Referrer": 40, "MIME Type": 18, "Download State": 13, "Danger Type": 20,
    "Download Size (bytes)": 14, "Received Bytes": 14, "Source": 20,
}

CHROMIUM_CORE_TRANSITIONS = {
    0: "LINK", 1: "TYPED", 2: "AUTO_BOOKMARK", 3: "AUTO_SUBFRAME",
    4: "MANUAL_SUBFRAME", 5: "GENERATED", 6: "START_PAGE", 7: "FORM_SUBMIT",
    8: "RELOAD", 9: "KEYWORD", 10: "KEYWORD_GENERATED",
}

CHROMIUM_QUALIFIERS = {
    0x00800000: "BLOCKED", 0x01000000: "FORWARD_BACK",
    0x02000000: "FROM_ADDRESS_BAR", 0x04000000: "HOME_PAGE",
    0x08000000: "FROM_API", 0x10000000: "CHAIN_START",
    0x20000000: "CHAIN_END", 0x40000000: "CLIENT_REDIRECT",
    0x80000000: "SERVER_REDIRECT",
}

FIREFOX_VISIT_TYPES = {
    1: "LINK", 2: "TYPED", 3: "BOOKMARK", 4: "EMBED",
    5: "REDIRECT_PERMANENT", 6: "REDIRECT_TEMPORARY", 7: "DOWNLOAD",
    8: "FRAMED_LINK", 9: "RELOAD",
}

CHROMIUM_DOWNLOAD_STATES = {0: "IN_PROGRESS", 1: "COMPLETE", 2: "CANCELLED", 3: "INTERRUPTED"}
FIREFOX_DOWNLOAD_STATES = {0: "DOWNLOADING", 1: "COMPLETE", 2: "FAILED", 3: "CANCELED", 4: "PAUSED"}
CHROMIUM_DANGER_TYPES = {
    0: "NOT_DANGEROUS", 1: "DANGEROUS_FILE", 2: "DANGEROUS_URL",
    3: "DANGEROUS_CONTENT", 4: "MAYBE_DANGEROUS_CONTENT", 5: "UNCOMMON_CONTENT",
    6: "USER_VALIDATED", 7: "DANGEROUS_HOST", 8: "POTENTIALLY_UNWANTED",
}

BROWSER_PATH_HINTS = [
    ("firefox", "Firefox"), ("mozilla", "Firefox"), ("chrome", "Chrome"),
    ("edge", "Edge"), ("brave", "Brave"), ("vivaldi", "Vivaldi"),
    ("opera", "Opera"), ("chromium", "Chromium"),
]


# --------------------------------------------------------------------------
# Time helpers (all output is UTC, ISO 8601 so it sorts correctly in Excel)
# --------------------------------------------------------------------------

def epoch_us_to_iso(epoch_us):
    """Epoch microseconds -> 'YYYY-MM-DD HH:MM:SS' UTC string."""
    if epoch_us in (None, ""):
        return ""
    try:
        dt = datetime.fromtimestamp(epoch_us / 1_000_000, tz=timezone.utc)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, OSError, OverflowError):
        return ""


def chromium_to_epoch_us(t):
    """Chromium/WebKit microseconds since 1601 -> epoch microseconds."""
    if not t:
        return None
    return t - CHROMIUM_EPOCH_OFFSET_US


def chromium_download_time_to_epoch_us(t):
    """Chromium download times: modern schemas store WebKit microseconds,
    very old schemas (pre-M26) stored plain epoch seconds. Heuristic on
    magnitude."""
    if not t:
        return None
    if t > 10**16:            # WebKit microseconds since 1601
        return t - CHROMIUM_EPOCH_OFFSET_US
    if t > 10**11:            # epoch milliseconds (defensive)
        return int(t * 1000)
    return int(t * 1_000_000)  # epoch seconds (legacy schema)


# --------------------------------------------------------------------------
# Generic helpers
# --------------------------------------------------------------------------

def blank_record():
    rec = {f: "" for f in FIELDNAMES}
    rec["_epoch_us"] = None  # internal, used for cross-source sorting; never exported
    return rec


def sha256_file(path, chunk=1024 * 1024):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for block in iter(lambda: f.read(chunk), b""):
            h.update(block)
    return h.hexdigest()


def table_exists(cursor, name):
    row = cursor.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)
    ).fetchone()
    return row is not None


def table_columns(cursor, name):
    return {row[1] for row in cursor.execute(f"PRAGMA table_info({name})")}


def decode_chromium_transition(transition):
    if transition is None:
        return "", "", ""
    core = transition & 0xFF
    core_name = CHROMIUM_CORE_TRANSITIONS.get(core, f"Unknown({core})")
    quals = [name for bit, name in CHROMIUM_QUALIFIERS.items() if transition & bit]
    return core, core_name, ", ".join(quals)


def detect_browser(cursor):
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = {row[0] for row in cursor.fetchall()}
    if {"urls", "visits"}.issubset(tables):
        return "chromium"
    if {"moz_places", "moz_historyvisits"}.issubset(tables):
        return "firefox"
    return None


# --------------------------------------------------------------------------
# Chromium parsing (Chrome / Edge / Brave / Vivaldi / Opera ...)
# --------------------------------------------------------------------------

def parse_chromium(cursor, errors):
    results = []

    # ---- Browsing history -------------------------------------------------
    try:
        vcols = table_columns(cursor, "visits")
        opener_expr = "visits.opener_visit" if "opener_visit" in vcols else "NULL"
        duration_expr = "visits.visit_duration" if "visit_duration" in vcols else "NULL"

        query = f"""
        SELECT visits.id, urls.url, urls.title, urls.visit_count,
               urls.last_visit_time, visits.visit_time, visits.from_visit,
               visits.transition, visits.segment_id, {opener_expr}, {duration_expr}
        FROM urls
        JOIN visits ON urls.id = visits.url
        ORDER BY visits.visit_time DESC;
        """
        for (visit_id, url, title, visit_count, last_visit, visit_time,
             from_visit, transition, segment_id, opener_visit,
             visit_duration) in cursor.execute(query):
            core_raw, core_name, quals = decode_chromium_transition(transition)
            rec = blank_record()
            rec.update({
                "Type": "Visit",
                "Visit ID": visit_id,
                "URL": url,
                "Title": title or "",
                "Visit Count": visit_count,
                "Visit Time (UTC)": epoch_us_to_iso(chromium_to_epoch_us(visit_time)),
                "Last Visit (UTC)": epoch_us_to_iso(chromium_to_epoch_us(last_visit)),
                "Visit Time Raw": visit_time,
                "Last Visit Raw": last_visit,
                "Visit Type": core_name,
                "Visit Type Raw": core_raw,
                "Transition Qualifiers": quals,
                "Transition Raw": transition,
                "From Visit ID": from_visit if from_visit else "",
                "Opener Visit ID": opener_visit if opener_visit else "",
                "Segment ID": segment_id if segment_id else "",
                "Visit Duration (s)": round(visit_duration / 1_000_000, 1) if visit_duration else "",
                "_epoch_us": chromium_to_epoch_us(visit_time),
            })
            results.append(rec)
    except sqlite3.Error as e:
        errors.append(f"Chromium history query failed: {e}")

    # ---- Downloads --------------------------------------------------------
    try:
        if table_exists(cursor, "downloads"):
            dcols = table_columns(cursor, "downloads")

            def col(name):
                return f"d.{name}" if name in dcols else "NULL"

            if table_exists(cursor, "downloads_url_chains"):
                chain_expr = ("(SELECT url FROM downloads_url_chains c "
                              "WHERE c.id = d.id ORDER BY c.chain_index DESC LIMIT 1)")
            else:
                chain_expr = "NULL"

            query = f"""
            SELECT {col('target_path')}, {col('current_path')}, {col('full_path')},
                   {col('start_time')}, {col('end_time')},
                   {col('received_bytes')}, {col('total_bytes')},
                   {col('state')}, {col('danger_type')}, {col('interrupt_reason')},
                   {col('mime_type')}, {col('tab_url')}, {col('tab_referrer_url')},
                   {col('url')}, {chain_expr}
            FROM downloads d;
            """
            for (target_path, current_path, full_path, start_time, end_time,
                 received_bytes, total_bytes, state, danger_type,
                 interrupt_reason, mime_type, tab_url, tab_referrer_url,
                 legacy_url, chain_url) in cursor.execute(query):
                start_epoch = chromium_download_time_to_epoch_us(start_time)
                end_epoch = chromium_download_time_to_epoch_us(end_time)
                rec = blank_record()
                rec.update({
                    "Type": "Download",
                    "URL": tab_url or legacy_url or "",
                    "Visit Time (UTC)": epoch_us_to_iso(start_epoch),
                    "End Time (UTC)": epoch_us_to_iso(end_epoch),
                    "Visit Time Raw": start_time if start_time else "",
                    "Download Path": target_path or current_path or full_path or "",
                    "Download URL": chain_url or legacy_url or "",
                    "Referrer": tab_referrer_url or "",
                    "MIME Type": mime_type or "",
                    "Download State": CHROMIUM_DOWNLOAD_STATES.get(state, state) if state is not None else "",
                    "Danger Type": CHROMIUM_DANGER_TYPES.get(danger_type, danger_type) if danger_type is not None else "",
                    "Interrupt Reason": interrupt_reason if interrupt_reason else "",
                    "Download Size (bytes)": total_bytes if total_bytes is not None else "",
                    "Received Bytes": received_bytes if received_bytes is not None else "",
                    "_epoch_us": start_epoch,
                })
                results.append(rec)
    except sqlite3.Error as e:
        errors.append(f"Chromium downloads query failed: {e}")

    # ---- Omnibox search terms --------------------------------------------
    try:
        if table_exists(cursor, "keyword_search_terms"):
            query = """
            SELECT k.term, u.url, u.last_visit_time
            FROM keyword_search_terms k
            JOIN urls u ON u.id = k.url_id;
            """
            for term, url, last_visit in cursor.execute(query):
                rec = blank_record()
                rec.update({
                    "Type": "Search Term",
                    "URL": url,
                    "Search Term": term,
                    "Last Visit (UTC)": epoch_us_to_iso(chromium_to_epoch_us(last_visit)),
                    "Last Visit Raw": last_visit,
                    "_epoch_us": chromium_to_epoch_us(last_visit),
                })
                results.append(rec)
    except sqlite3.Error as e:
        errors.append(f"Chromium keyword_search_terms query failed: {e}")

    return results


# --------------------------------------------------------------------------
# Firefox parsing
# --------------------------------------------------------------------------

def parse_firefox(cursor, errors):
    results = []

    # ---- Browsing history -------------------------------------------------
    try:
        query = """
        SELECT v.id, p.url, p.title, p.visit_count, p.last_visit_date,
               v.visit_date, v.visit_type, v.from_visit
        FROM moz_places p
        JOIN moz_historyvisits v ON p.id = v.place_id
        ORDER BY v.visit_date DESC;
        """
        for (visit_id, url, title, visit_count, last_visit, visit_time,
             visit_type, from_visit) in cursor.execute(query):
            rec = blank_record()
            rec.update({
                "Type": "Visit",
                "Visit ID": visit_id,
                "URL": url,
                "Title": title or "",
                "Visit Count": visit_count,
                "Visit Time (UTC)": epoch_us_to_iso(visit_time),
                "Last Visit (UTC)": epoch_us_to_iso(last_visit),
                "Visit Time Raw": visit_time,
                "Last Visit Raw": last_visit,
                "Visit Type": FIREFOX_VISIT_TYPES.get(visit_type, f"Unknown({visit_type})"),
                "Visit Type Raw": visit_type,
                "From Visit ID": from_visit if from_visit else "",
                "_epoch_us": visit_time,
            })
            results.append(rec)
    except sqlite3.Error as e:
        errors.append(f"Firefox history query failed: {e}")

    # ---- Legacy downloads table (very old Firefox) ------------------------
    try:
        if table_exists(cursor, "moz_downloads"):
            for target, start_time, total_bytes, source in cursor.execute(
                "SELECT target, startTime, totalBytes, source FROM moz_downloads;"
            ):
                rec = blank_record()
                rec.update({
                    "Type": "Download",
                    "URL": source or "",
                    "Download URL": source or "",
                    "Visit Time (UTC)": epoch_us_to_iso(start_time),
                    "Visit Time Raw": start_time if start_time else "",
                    "Download Path": target or "",
                    "Download Size (bytes)": total_bytes if total_bytes is not None else "",
                    "_epoch_us": start_time,
                })
                results.append(rec)
    except sqlite3.Error as e:
        errors.append(f"Firefox moz_downloads query failed: {e}")

    # ---- Modern downloads (moz_annos annotations) -------------------------
    # NOTE: metaData is LEFT JOINed via subquery so downloads without a
    # metaData annotation still appear (with blank size/time). metaData's
    # endTime is epoch *milliseconds* and is the completion time -- Firefox
    # does not persist a start time here.
    try:
        if table_exists(cursor, "moz_annos"):
            query = """
            SELECT p.url, d_file.content AS target_path, meta.content AS meta_json
            FROM moz_places p
            JOIN moz_annos d_file ON d_file.place_id = p.id
            JOIN moz_anno_attributes a_file
              ON a_file.id = d_file.anno_attribute_id
             AND a_file.name = 'downloads/destinationFileURI'
            LEFT JOIN (
                SELECT ann.place_id, ann.content
                FROM moz_annos ann
                JOIN moz_anno_attributes attr ON attr.id = ann.anno_attribute_id
                WHERE attr.name = 'downloads/metaData'
            ) meta ON meta.place_id = p.id;
            """
            for source_url, target_path, meta_json in cursor.execute(query):
                total_bytes = ""
                end_epoch_us = None
                state = ""
                if meta_json:
                    try:
                        meta = json.loads(meta_json)
                        total_bytes = meta.get("fileSize", "")
                        end_ms = meta.get("endTime")
                        if end_ms:
                            end_epoch_us = int(end_ms) * 1000
                        raw_state = meta.get("state")
                        if raw_state is not None:
                            state = FIREFOX_DOWNLOAD_STATES.get(raw_state, raw_state)
                    except (ValueError, TypeError):
                        pass
                rec = blank_record()
                rec.update({
                    "Type": "Download",
                    "URL": source_url or "",
                    "Download URL": source_url or "",
                    "End Time (UTC)": epoch_us_to_iso(end_epoch_us),
                    "Visit Time Raw": end_epoch_us if end_epoch_us else "",
                    "Download Path": target_path.replace("file://", "") if target_path else "",
                    "Download Size (bytes)": total_bytes,
                    "Download State": state,
                    "_epoch_us": end_epoch_us,
                })
                results.append(rec)
    except sqlite3.Error as e:
        errors.append(f"Firefox moz_annos downloads query failed: {e}")

    # ---- Awesome-bar typed input ------------------------------------------
    try:
        if table_exists(cursor, "moz_inputhistory"):
            query = """
            SELECT i.input, p.url, p.last_visit_date
            FROM moz_inputhistory i
            JOIN moz_places p ON p.id = i.place_id;
            """
            for typed, url, last_visit in cursor.execute(query):
                rec = blank_record()
                rec.update({
                    "Type": "Typed Input",
                    "URL": url,
                    "Search Term": typed,
                    "Last Visit (UTC)": epoch_us_to_iso(last_visit),
                    "Last Visit Raw": last_visit if last_visit else "",
                    "_epoch_us": last_visit,
                })
                results.append(rec)
    except sqlite3.Error as e:
        errors.append(f"Firefox moz_inputhistory query failed: {e}")

    return results


# --------------------------------------------------------------------------
# Per-source processing
# --------------------------------------------------------------------------

def process_db(db_path):
    """Copy DB (+ WAL/SHM sidecars) to a temp dir, parse, clean up."""
    info = {
        "path": os.path.abspath(db_path),
        "sha256": "",
        "wal": False,
        "shm": False,
        "browser": None,
        "records": [],
        "errors": [],
    }

    if not os.path.isfile(db_path):
        info["errors"].append("File not found")
        return info

    try:
        info["sha256"] = sha256_file(db_path)
    except OSError as e:
        info["errors"].append(f"Could not hash source: {e}")

    tmpdir = tempfile.mkdtemp(prefix="bh_")
    conn = None
    try:
        temp_db = os.path.join(tmpdir, "history.db")
        try:
            shutil.copy2(db_path, temp_db)
        except OSError as e:
            info["errors"].append(
                f"Could not copy database ({e}). If the browser is running, "
                "collect via VSS shadow copy or your EDR's raw file read."
            )
            return info

        # WAL/SHM sidecars hold recent, un-checkpointed activity. Copy them so
        # SQLite replays the WAL when we open our copy.
        for suffix, key in (("-wal", "wal"), ("-shm", "shm")):
            side = db_path + suffix
            if os.path.exists(side):
                info[key] = True
                try:
                    shutil.copy2(side, temp_db + suffix)
                except OSError as e:
                    info["errors"].append(f"Could not copy {suffix} sidecar: {e}")

        conn = sqlite3.connect(temp_db)
        cursor = conn.cursor()

        info["browser"] = detect_browser(cursor)
        if info["browser"] is None:
            info["errors"].append("Unknown or unsupported browser database schema")
            return info

        if info["browser"] == "chromium":
            info["records"] = parse_chromium(cursor, info["errors"])
        else:
            info["records"] = parse_firefox(cursor, info["errors"])

        info["records"].sort(
            key=lambda r: r["_epoch_us"] if r["_epoch_us"] is not None else -1,
            reverse=True,
        )
    except sqlite3.Error as e:
        info["errors"].append(f"SQLite error: {e}")
    finally:
        if conn is not None:
            conn.close()
        shutil.rmtree(tmpdir, ignore_errors=True)

    return info


def counts(info):
    visits = sum(1 for r in info["records"] if r["Type"] == "Visit")
    downloads = sum(1 for r in info["records"] if r["Type"] == "Download")
    terms = sum(1 for r in info["records"] if r["Type"] in ("Search Term", "Typed Input"))
    return visits, downloads, terms


# --------------------------------------------------------------------------
# Input discovery & sheet naming
# --------------------------------------------------------------------------

def discover_dbs(paths):
    """Expand any directories into contained History / places.sqlite files."""
    found = []
    for p in paths:
        if os.path.isdir(p):
            hits = []
            for root, _dirs, files in os.walk(p):
                for f in files:
                    if f.lower() in ("history", "places.sqlite"):
                        hits.append(os.path.join(root, f))
            if not hits:
                print(f"[!] No History / places.sqlite files found under: {p}")
            found.extend(sorted(hits))
        else:
            found.append(p)
    return list(dict.fromkeys(os.path.abspath(f) for f in found))


def derive_label(path, browser_type, taken):
    """Best-effort human-readable sheet name from the path, e.g. Chrome-Default,
    Edge-Profile 1, Firefox-Default. Excel limits: 31 chars, no []:*?/\\ ."""
    parts = [p for p in re.split(r"[\\/]+", path) if p]
    lower = [p.lower() for p in parts]

    browser_name = None
    for hint, name in BROWSER_PATH_HINTS:
        if any(hint in part for part in lower):
            browser_name = name
            break
    if browser_name is None:
        browser_name = (browser_type or "unknown").capitalize()

    profile = None
    for part, low in zip(parts, lower):
        if low == "default" or ".default" in low:
            profile = "Default"
            break
        if low.startswith("profile") and low != "profiles":
            profile = part
            break

    label = f"{browser_name}-{profile}" if profile else browser_name
    label = re.sub(r"[\[\]:*?/\\]", "-", label).strip() or "Source"
    label = label[:28]

    candidate, n = label, 1
    while candidate in taken:
        n += 1
        candidate = f"{label}-{n}"[:31]
    return candidate


# --------------------------------------------------------------------------
# XLSX output
# --------------------------------------------------------------------------

def write_xlsx(sources, combined, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    def clean(v):
        # Strip XML-illegal control chars (titles/URLs are untrusted content
        # and can contain them); U+FFFD keeps the tampering visible.
        if isinstance(v, str):
            return ILLEGAL_CHARACTERS_RE.sub("\ufffd", v)
        return "" if v is None else v

    def append_row(ws, values):
        cleaned = [clean(v) for v in values]
        ws.append(cleaned)
        # openpyxl treats strings starting with '=' as formulas; force them
        # back to text so hostile page titles can't execute in Excel.
        row_idx = ws.max_row
        for col_idx, v in enumerate(cleaned, start=1):
            if isinstance(v, str) and v.startswith("="):
                ws.cell(row=row_idx, column=col_idx).data_type = "s"

    def style_table(ws, header_row=1, n_cols=None, widths_for=None):
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(n_cols)}{ws.max_row}"
        )
        for idx, name in enumerate(widths_for, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(name, 13)

    wb = Workbook()

    # ---- Summary ----------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    append_row(ws, [f"Browser history export - generated {generated}"])
    ws["A1"].font = Font(bold=True, size=12)
    append_row(ws, [])
    summary_fields = ["Sheet", "Detected", "Visits", "Downloads",
                      "Search/Typed Terms", "WAL Sidecar", "SHM Sidecar",
                      "SHA-256 (source)", "Source Path", "Errors"]
    append_row(ws, summary_fields)
    for s in sources:
        v, d, t = counts(s)
        append_row(ws, [
            s["label"],
            (s["browser"] or "unknown").capitalize(),
            v, d, t,
            "Yes" if s["wal"] else "No",
            "Yes" if s["shm"] else "No",
            s["sha256"],
            s["path"],
            "; ".join(s["errors"]),
        ])
    for cell in ws[3]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A4"
    for idx, name in enumerate(summary_fields, start=1):
        width = {"Sheet": 22, "Detected": 12, "SHA-256 (source)": 66,
                 "Source Path": 70, "Errors": 50}.get(name, 14)
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ---- Combined Timeline ------------------------------------------------
    ws = wb.create_sheet("Combined Timeline")
    combined_fields = ["Source"] + FIELDNAMES
    append_row(ws, combined_fields)
    rows = combined
    if len(rows) > EXCEL_MAX_ROWS - 1:
        print(f"[!] Combined Timeline truncated to {EXCEL_MAX_ROWS - 1:,} rows "
              f"(had {len(rows):,}). Per-source sheets are unaffected.")
        rows = rows[: EXCEL_MAX_ROWS - 1]
    for label, rec in rows:
        append_row(ws, [label] + [rec.get(f, "") for f in FIELDNAMES])
    style_table(ws, n_cols=len(combined_fields), widths_for=combined_fields)

    # ---- One sheet per source --------------------------------------------
    for s in sources:
        ws = wb.create_sheet(s["label"])
        append_row(ws, FIELDNAMES)
        for rec in s["records"]:
            append_row(ws, [rec.get(f, "") for f in FIELDNAMES])
        style_table(ws, n_cols=len(FIELDNAMES), widths_for=FIELDNAMES)

    wb.save(out_path)
    print(f"[+] Export complete: {out_path}")


# --------------------------------------------------------------------------
# CSV fallback (no openpyxl, or -o something.csv)
# --------------------------------------------------------------------------

def write_csv_fallback(sources, combined, out_path):
    stem = os.path.splitext(out_path)[0]

    def guard(v):
        # CSV opened in Excel will execute =, +, -, @ prefixed cells.
        s = "" if v is None else str(v)
        return "'" + s if s[:1] in ("=", "+", "-", "@") else s

    written = []
    for s in sources:
        path = f"{stem}_{s['label']}.csv"
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(FIELDNAMES)
            for rec in s["records"]:
                w.writerow([guard(rec.get(fn, "")) for fn in FIELDNAMES])
        written.append(path)

    path = f"{stem}_Combined.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Source"] + FIELDNAMES)
        for label, rec in combined:
            w.writerow([guard(label)] + [guard(rec.get(fn, "")) for fn in FIELDNAMES])
    written.append(path)

    print("[+] CSV export complete:")
    for p in written:
        print(f"      {p}")


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description="Parse one or more Chromium History / Firefox places.sqlite "
                    "databases into a single Excel workbook (one sheet per "
                    "source, plus a combined cross-browser timeline).",
    )
    ap.add_argument("paths", nargs="+",
                    help="History/places.sqlite file(s), or directories to scan for them")
    ap.add_argument("-o", "--output", default="browser_history_export.xlsx",
                    help="Output .xlsx path (use a .csv extension to force CSV mode)")
    args = ap.parse_args()

    db_paths = discover_dbs(args.paths)
    if not db_paths:
        print("[!] No history databases to process.")
        sys.exit(1)

    sources = []
    taken = set()
    for p in db_paths:
        info = process_db(p)
        info["label"] = derive_label(p, info["browser"], taken)
        taken.add(info["label"])
        sources.append(info)

        v, d, t = counts(info)
        if info["browser"]:
            wal_note = " (+WAL)" if info["wal"] else ""
            print(f"[+] {info['label']}: {info['browser'].capitalize()}{wal_note} - "
                  f"{v} visits, {d} downloads, {t} search/typed terms")
        for err in info["errors"]:
            print(f"    [!] {info['label']}: {err}")

    if not any(s["records"] for s in sources):
        print("[!] No history or download data found in any source.")

    combined = [(s["label"], r) for s in sources for r in s["records"]]
    combined.sort(
        key=lambda item: item[1]["_epoch_us"] if item[1]["_epoch_us"] is not None else -1,
        reverse=True,
    )

    force_csv = args.output.lower().endswith(".csv")
    if not force_csv:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            print("[!] openpyxl not installed (pip install openpyxl) - "
                  "falling back to CSV output.")
            force_csv = True

    if force_csv:
        write_csv_fallback(sources, combined, args.output)
    else:
        write_xlsx(sources, combined, args.output)


if __name__ == "__main__":
    main()
